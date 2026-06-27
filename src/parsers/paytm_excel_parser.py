"""Parser for Paytm UPI statement .xlsx exports.

Paytm is a UPI front-end on top of the user's bank accounts (ICICI, HDFC, PNB,
SBI) and credit cards. Each Paytm payment therefore *also* exists in the
underlying bank statement IF that bank's statement is loaded. The crucial field
is the **UPI Ref No.** (NPCI RRN), which is identical across Paytm and the bank
narration — the merge step (paytm_merge.py) uses it to avoid double counting.

Observed layout (modern .xlsx, read with openpyxl):

    Sheet 'Summary'                — totals + per-account breakdown (skipped here)
    Sheet 'Passbook Payment History' — the transactions, header on row 0:
        Date | Time | Transaction Details | Other Transaction Details |
        Your Account | Amount | UPI Ref No. | Order ID | Remarks | Tags | Comment

    * Date is dd/mm/yyyy; Amount is a signed string like '-1,174.00' / '+1,500.00'.
    * 'Your Account' is the funding source, e.g. 'Punjab National Bank - 77',
      'ICICI Bank - 52', 'HDFC Bank Rupay Credit Card'.
    * 'Other Transaction Details' holds the counterparty VPA (e.g. 8617663869@paytm).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from src.models.transaction_schema import PARSER_OUTPUT_COLUMNS
from src.parsers.base import BaseParser, ParseResult
from src.utils.dates import parse_date

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

_SHEET = "Passbook Payment History"

# Map Paytm "Your Account" labels to short, stable source-bank labels.
_BANK_MAP = [
    ("punjab national", "PNB (via Paytm)"),
    ("state bank", "SBI (via Paytm)"),
    ("icici bank rupay credit", "ICICI-CC (via Paytm)"),
    ("hdfc bank rupay credit", "HDFC-CC (via Paytm)"),
    ("axis bank rupay credit", "Axis-CC (via Paytm)"),
    ("axis bank credit", "Axis-CC (via Paytm)"),
    ("rupay credit", "CC (via Paytm)"),
    ("credit card", "CC (via Paytm)"),
    ("icici", "ICICI (via Paytm)"),
    ("hdfc", "HDFC (via Paytm)"),
    ("upi lite", "Paytm-Lite"),
]


def _map_bank(account: str) -> str:
    a = (account or "").lower()
    for needle, label in _BANK_MAP:
        if needle in a:
            return label
    return "Paytm"


def _amount(value) -> tuple[float, float]:
    """Return (debit, credit) from a signed Paytm amount string/number."""
    if value is None:
        return 0.0, 0.0
    s = str(value).replace(",", "").replace("₹", "").strip()
    try:
        n = float(s)
    except ValueError:
        return 0.0, 0.0
    return (abs(n), 0.0) if n < 0 else (0.0, abs(n))


class PaytmExcelParser(BaseParser):
    bank_name = "Paytm"
    supported_extensions = [".xlsx"]

    def can_parse(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions:
            return False
        name = path.name.lower()
        folder = path.parent.name.lower()
        return "paytm" in folder or "paytm" in name

    def parse(self, path: Path) -> ParseResult:
        if openpyxl is None:
            return self._empty_result("openpyxl is required to read Paytm .xlsx files.", path)
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            return self._empty_result(f"Could not open Paytm workbook: {exc}", path)

        if _SHEET not in wb.sheetnames:
            return self._empty_result(f"Paytm sheet '{_SHEET}' not found.", path)
        ws = wb[_SHEET]

        rows, warnings = self._read_rows(ws, path)
        df = pd.DataFrame(rows, columns=PARSER_OUTPUT_COLUMNS)
        return ParseResult(
            transactions=df,
            warnings=warnings,
            metadata={"sheet_name": _SHEET, "rows_extracted": len(df),
                      "stated_date_from": None, "stated_date_to": None},
        )

    def _read_rows(self, ws, path: Path):
        rows, warnings = [], []
        seen_header = False
        for r in ws.iter_rows(values_only=True):
            if r is None:
                continue
            first = str(r[0] or "").strip()
            if not seen_header:
                if first.lower() == "date":
                    seen_header = True
                continue
            if not first:  # blank trailing row
                continue
            date = parse_date(first)
            if not date:
                continue

            details = str(r[2] or "").strip()                 # "Paid to X" / "Money sent to X"
            other = str(r[3] or "").strip()                   # counterparty VPA
            account = str(r[4] or "").strip()                 # funding source
            debit, credit = _amount(r[5])
            ref = str(r[6] or "").strip()
            remarks = str(r[8] or "").strip()

            narration = " | ".join(p for p in [details, other, f"via {account}" if account else "", remarks] if p)
            rows.append({
                "source_bank": _map_bank(account),
                "source_file": path.name,
                "source_folder": path.parent.name,
                "source_sheet": _SHEET,
                "source_row_number": ws._current_row if hasattr(ws, "_current_row") else None,
                "source_parser": "paytm_excel_parser",
                "source_format": "xlsx",
                "transaction_date": date,
                "value_date": date,
                "description": details or other,
                "raw_description": narration,
                "reference_number": ref,        # NPCI RRN — used for dedup
                "cheque_number": "",
                "debit": debit,
                "credit": credit,
                "balance": None,                # Paytm statements carry no running balance
            })

        if not rows:
            warnings.append("No Paytm transactions were extracted.")
        else:
            dts = pd.to_datetime([x["transaction_date"] for x in rows], errors="coerce")
            earliest = pd.Series(dts).min()
            if pd.notna(earliest) and earliest.year >= 2024:
                warnings.append(
                    "Paytm history starts "
                    f"{earliest.date()}; payments before 2024 are not available "
                    "(acceptable — older Benazir/other payments may be incomplete)."
                )
        return rows, warnings
